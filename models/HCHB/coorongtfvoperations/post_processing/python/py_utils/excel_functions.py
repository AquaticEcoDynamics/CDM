"""
Functions to export Physical and Biogeochemical WQ Statistics to formatted
excel tables. 
Designed for CoorongBGC Project, A10780.
A.Waterhouse
"""

from typing import Union
import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.worksheet.cell_range import CellRange
from pathlib import Path
import pandas as pd
import numpy as np
import re
from matplotlib import cm
from matplotlib.colors import rgb2hex


def get_sheetnames_xlsx(filepath):
    wb = openpyxl.load_workbook(filepath, read_only=True, keep_links=False)
    return wb.sheetnames

def add_border_to_range(
    cells, border_style="thin", font_size=9, font_name="Arial", font_b=False
):
    """Helper script to add borders using openpyxl

    Args:
        ws ([type]): openpyxl worksheet
        cell_range (str): Cell range to add borders to (e.g. "A1:B4")
        border_style (str, optional): Border style. Defaults to 'thin'.
        font_size (int, optional): Font size. Defaults to 9.
        font_name (str, optional): Font Name. Defaults to "Arial".
        font_b (bool, optional): Bold bool. Defaults to False.
    """

    font = Font(size=font_size, name=font_name, b=font_b)
    align = Alignment(horizontal="left", vertical="center")
    border = Border(
        left=Side(border_style=border_style, color="000000"),
        right=Side(border_style=border_style, color="000000"),
        top=Side(border_style=border_style, color="000000"),
        bottom=Side(border_style=border_style, color="000000"),
    )

    rows = [rows for rows in cells]
    flattened = [item for sublist in rows for item in sublist]
    [
        (
            setattr(cell, "border", border),
            setattr(cell, "font", font),
            setattr(cell, "alignment", align),
        )
        for cell in flattened
    ]


def cellstr(
    x1: int, y1: int, x2: Union[int, None] = None, y2: Union[int, None] = None
) -> str:
    """Generate excel style string cell range
    E.g. cellstr(1,1,1,4) = A1:A4

    Args:
        x1 (int): Column Start
        y1 (int): Row Start
        x2 (int): Column End
        y2 (int): Row End

    Returns:
        str: Excel style cell range string
    """
    if x2 == None:
        x2 = x1
    if y2 == None:
        y2 = y1
    return str(CellRange(min_col=x1, max_col=x2, min_row=y1, max_row=y2))


def write_excel_document(
    excel_file_path: Path,
    results: dict,
    variables: dict,
    locations: dict,
    seasons: dict,
):
    """Function to write statistic excel tables for TUFLOWFV results.

    Args:
        excel_file_path (Path): Path object to excel output file
        results (dict): results loaded in a dictionary.
        variables (dict): Variables to output in a dict. (Key: TUFLOWFV Short name, Value: Label name (e.g. Surface Elevation))
        stations (dict): Stations to process. (Key: Name as in the CSV/Dict file, Value: Label name (e.g. Parnka Point))
        periods (dict): Periods/seasons. (Key: Season Name, Value: tuple of start/end (pd.Timestamp, pd.Timestamp)))

    Requires openpyxl
    """

    stats = {"Mean": np.mean, "Median": np.median, "Min": np.min, "Max": np.max}

    # ToDO: Fix the bad coding - hardcoded dictionary key while in a hurry
    scn_ids = results.keys()
    if "SC01" in scn_ids:
        base = "SC01"
    elif "SC14" in scn_ids:
        base = "SC14"
    else:
        base = "Basecase"

    # We're going to make a nested pandas df so we need to initialise the multi-index hierachy first!
    iters = [stats.keys(), scn_ids]
    index = pd.MultiIndex.from_product(iters)

    iters = [seasons.keys(), locations.values(), ["Value", "Δ"]]
    columns = pd.MultiIndex.from_product(iters)

    writer = pd.ExcelWriter(excel_file_path, engine="openpyxl")
    excel_vars = {}
    for var, vname in variables.items():
        df = pd.DataFrame(index=index, columns=columns).sort_index()
        for season, (time_start, time_end) in seasons.items():
            for c, (loc_id, loc) in enumerate(locations.items()):
                for stat, fn in stats.items():
                    for scn_id, df_scn in results.items():
                        col = [x for x in df_scn.columns if loc_id in x if var in x][0]

                        # Filter dataframe by time!
                        df_scn = df_scn[
                            (df_scn["TIME"] >= time_start) & (df_scn["TIME"] < time_end)
                        ]

                        df.loc[(stat, scn_id), (season, loc, "Value")] = fn(df_scn[col])

                    df.loc[(stat,), (season, loc, "Δ")] = (
                        df.loc[(stat,), (season, loc, "Value")]
                        - df.loc[(stat, base), (season, loc, "Value")]
                    ).values

        dims = df.shape
        df = df.loc[index, columns]
        df.to_excel(writer, sheet_name=var)
        writer.save()
        excel_vars[var] = (vname, df)

    format_excel_output(excel_file_path, excel_vars)


def format_excel_output(excel_file_path, variables):
    # Load up the excel format sheet
    excelFormat = ExcelFormat()

    hdr_font = Font(b=True, color="000000", size=10, name="Arial")
    num_font = Font(b=False, color="000000", size=10, name="Arial")
    alng = Alignment(horizontal="center", vertical="center")

    wb = openpyxl.load_workbook(filename=excel_file_path)
    start_col = 3
    start_row = 5

    for variable, (var_full_name, df) in variables.items():
        ws = wb[variable]  # Choose current worksheet by variable

        # Get dims and vars
        dims = df.shape
        end_row = dims[0] + 4
        index_vars = df.index.levels[0]
        column_vars = df.columns.levels[-1]

        # Count of columns
        lev = df.columns.levshape

        # Bit of nasty gymnastics to unmerge and remerge header cols
        season_shp = int(dims[1] / lev[0])
        loc_shp = int(season_shp / lev[1])
        col = start_col
        for col_idx in range(lev[0]):
            ws.unmerge_cells(cellstr(col, 1, col + season_shp - 1, 1))

            # Unmerge the second row columns (e.g multiple locs)
            for n in range(lev[1]):
                ws.unmerge_cells(cellstr(col + n * loc_shp, 2, col + (n + 1) * loc_shp - 1, 2))
            
            ws.move_range(cellstr(col, 1, col + season_shp - 1, 3), rows=1, cols=0)
            ws.merge_cells(cellstr(col, 2, col + season_shp - 1, 2))

            # Re-merge the now third row columns (e.g multiple locs)
            for n in range(lev[1]):
                ws.merge_cells(cellstr(col + n * loc_shp, 3, col + (n + 1) * loc_shp - 1, 3))
            col += season_shp

        # Add borders now everything is nicely merged
        add_border_to_range(ws[cellstr(1, 4, dims[1] + start_col - 1, end_row)], "thin")
        add_border_to_range(
            ws[cellstr(start_col, 1, dims[1] + start_col - 1, 4)], "thin"
        )

        # Add the variable name
        ws[cellstr(start_col, 1)] = var_full_name
        ws.merge_cells(cellstr(start_col, 1, start_col + dims[1] - 1))
        ws[cellstr(start_col, 1)].font = hdr_font
        ws[cellstr(start_col, 1)].alignment = alng

        ws["A4"] = "Statistic"
        ws["B4"] = "Scenario"

        # Iterate through columns in the dataframe and apply styling
        col = start_col
        for col_idx in df.columns:
            # Col var (e.g. Value, Delta, etc.)
            col_var = col_idx[2]

            # Header Row Formatting
            for row in range(2, 5):
                cell = cellstr(col, row)
                ws[cell].font = hdr_font
                ws[cell].alignment = alng

            # Width formatting
            ws.column_dimensions[re.match("\D*", cellstr(col, 1))[0]].width = 12

            # Add formatting as specified in format csv
            row = start_row
            for row_idx in df[col_idx].index:
                row_var = row_idx[0]
                cell = cellstr(col, row)

                # Apply standard font
                ws[cell].font = num_font

                # Apply variable specific formatting
                fmt = excelFormat.get_format(variable, row_var, col_var)
                ws[cell].number_format = fmt["fmt"]
                if fmt["rule"]:
                    ws.conditional_formatting.add(cell, fmt["rule"])
                row += 1
            col += 1

        # Fix index columns
        ws["B1"].style = "Normal"
        ws["B2"].style = "Normal"
        ws["B3"].style = "Normal"
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 12

        for col in [1, 2]:
            for row in range(4, end_row + 1):
                ws[cellstr(col, row)].font = hdr_font
                ws[cellstr(col, row)].alignment = alng

        # Finally, freeze the pane
        ws.freeze_panes = "C5"

    wb.save(filename=excel_file_path)


class ExcelFormat:
    """Class to get read variable format from csv"""

    def __init__(self):
        self.df = pd.read_csv(
            Path(__file__).parent.absolute() / "excel_table_format.csv",
            index_col=[0],
            dtype=str,
        )

    def get_format(self, variable, index_filter, col_filter):
        df = self.df
        clrs = {"red": "F8696B", "green": "63BE7B", "white": "FFFFFF"}

        # Defaults
        fmt = {"fmt": "0.0", "rule": None}

        msk = (
            (df.index == variable)
            & ((df["IndexVar"].values == index_filter) | (df["IndexVar"] == "ANY"))
            & ((df["ColumnVar"] == col_filter) | (df["ColumnVar"] == "ANY"))
        )
        if sum(msk) > 0:
            row = df[msk].iloc[0, :].to_dict()
            fmt["fmt"] = row["Format"]
        
            if row["Colour"] == "TRUE":
                if "cmap" in str(row["LowerColour"]):
                    cmap_name = row["LowerColour"].split("_")[1]
                    cmap = cm.get_cmap("GnBu", 3)
                    start_colour = rgb2hex(cmap(0))[1:]
                    mid_colour = rgb2hex(cmap(1))[1:]
                    end_colour = rgb2hex(cmap(2))[1:]

                elif "#" in str(row["LowerColour"]):
                    start_colour = row["LowerColour"][1:]
                    mid_colour = row["MidColour"][1:]
                    end_colour = row["UpperColour"][1:]

                else:
                    start_colour = clrs[row["LowerColour"]]
                    mid_colour = clrs[row["MidColour"]]
                    end_colour = clrs[row["UpperColour"]]
                
                fmt["rule"] = openpyxl.formatting.rule.ColorScaleRule(
                    start_type="num",
                    start_value=float(row["Lower"]),
                    start_color=start_colour,
                    mid_type="num",
                    mid_value=float(row["Mid"]),
                    mid_color=mid_colour,
                    end_type="num",
                    end_value=float(row["Upper"]),
                    end_color=end_colour,
                )
            else:
                fmt["rule"] = None
        return fmt
